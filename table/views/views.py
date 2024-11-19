import os
import zipfile
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
import json
from openpyxl.reader.excel import load_workbook
from django.http import HttpResponseRedirect, HttpResponseBadRequest
import unicodedata
from django.db.models import Count
from django.contrib.auth.decorators import login_required

from .planilha_distribuicao import *
from .planilha_docentes import *
from .salvar_modificacoes import *
from .preferencias_upload import *

@login_required
def index(request, semestre, ano):
    # cria listas com referências diferentes
    # elas serão iteradas no index.html para carregar os dados na planilha editável
    # Necessita de modularização

    tbl_vls = [[""] * 10 for _ in range(8)]
    tbl_vls.insert(3, ["", ""])

    tbl_extra = [[""] * 10 for _ in range(4)]

    query_smt = [7, 8] if semestre in [7, 8] else [semestre]
    smt_ano = "I" if semestre % 2 else "P"
    vls_turmas = Turma.objects.filter(
        Q(CoDisc__SemestreIdeal__in=query_smt, Ano=ano, SemestreAno=smt_ano, Eextra="N") |
        Q(Ano=ano, Eextra="S", SemestreAno=smt_ano, semestre_extra=semestre)
    )

    turmas_rp_db = Turma.objects.filter(
        Q(CoDisc="ACH0042", Ano=ano, SemestreAno=smt_ano, Eextra="N") |
        Q(CoDisc="ACH0042", Ano=ano, Eextra="S", SemestreAno=smt_ano, semestre_extra=semestre)
    )

    for tur_materia in vls_turmas:
        cod_disc = tur_materia.CoDisc
        if str(cod_disc) == "ACH0042":
            prof = []
            for turma_rp in turmas_rp_db:
                if(turma_rp.CodTurma == tur_materia.CodTurma):
                    prof.append(turma_rp.NroUSP.Apelido)
            prof = " / ".join(prof)
        else: prof = tur_materia.NroUSP.Apelido
        t_nro = int(tur_materia.CodTurma)
        t_extra = tur_materia.Eextra

        # faz a uma consulta na relação N:N
        dias_aulas = tur_materia.dia_set.all()

        for dia in dias_aulas:

            col_tbl = int(dia.DiaSemana)
            row_tbl = int(dia.Horario)

            if row_tbl in (5, 7) and t_nro == 94:
                row_tbl += 1

            # caso do vespertino 1 com duas linhas
            # 33 é um número arbitrário
            if t_nro == 33:
                row_tbl += 1

            if t_extra == "S" and row_tbl not in (2, 3, 4):
                if row_tbl == 5:
                    row_tbl = 2
                elif row_tbl == 7:
                    row_tbl = 3
                atribuir_tbl_values(tbl_extra, cod_disc, row_tbl, col_tbl, prof)
            else:
                atribuir_tbl_values(tbl_vls, cod_disc, row_tbl, col_tbl, prof)

    rest_turno = {"manha": 0, "tarde": 22, "noite": 48}
    dia_sem = {"segunda": 0, "terca": 2, "quarta": 4, "quinta": 6, "sexta": 8}
    profs_objs = Professor.objects.all()
    restricoes_profs = {}
    impedimentos_totais = {}

    # Decide quais restrições serão carregadas
    s_rest = "1" if semestre % 2 else "2"

    auto_profs = {}  # para o autocomplete do nome do professor com apelido
    detalhes_profs = {}
    for prof_obj in profs_objs:
        # Já carrega dados necessários para as sugestões e detalhes do professor
        nome = unicodedata.normalize("NFD", prof_obj.NomeProf.lower())
        auto_profs[prof_obj.NomeProf] = prof_obj.Apelido
        consideracao = prof_obj.consideracao1 if semestre % 2 else prof_obj.consideracao2
        detalhes_profs[nome] = [prof_obj.NomeProf, prof_obj.Apelido, prof_obj.pos_doc, prof_obj.pref_optativas,
                                consideracao]

        # tentar melhorar desempenho da linha abaixo
        restricoes = prof_obj.restricao_set.filter(semestre=s_rest)

        restricoes_profs[str(prof_obj.Apelido)] = []
        impedimentos_totais[str(prof_obj.Apelido)] = []

        for rest_prof in restricoes:
            list_rest_indice = []
            indice = dia_sem[rest_prof.dia] + rest_turno[rest_prof.periodo]
            if rest_prof.periodo == "tarde" and rest_prof.dia == "segunda":
                list_rest_indice = [indice, 23, 33, 34, 35, 36]
            elif rest_prof.periodo == "tarde":
                list_rest_indice = [indice, indice + 1, indice + 13, indice + 14]
            elif rest_prof.periodo == "noite":
                indice = indice - 2
                list_rest_indice = [
                    indice,
                    indice + 1,
                    indice + 11,
                    indice + 12,
                    indice + 22,
                    indice + 23,
                    indice + 33,
                    indice + 34,
                ]
            else:
                list_rest_indice = [indice, indice + 1, indice + 11, indice + 12]

            if str(prof_obj.Apelido) in restricoes_profs:
                restricoes_profs[str(prof_obj.Apelido)] += list_rest_indice
            else:
                restricoes_profs[str(prof_obj.Apelido)] = list_rest_indice

            # impedimento total
            if rest_prof.impedimento:
                impedimentos_totais[str(prof_obj.Apelido)] = list_rest_indice

    if ano == AnoAberto.objects.get(id=1).Ano:
        discs = Disciplina.objects.filter(ativa=True)
    else:
        discs = Disciplina.objects.all()

    cods_tbl_hr = {}
    cods_tbl_hr_ext = {}  # códigos para preencher a tabela de matérias que não são do semestre selecionado
    mtr_auto_nome = {}  # auxilia no autocomplete para encontrar o código da turma extra

    cod_mtr_sugestao = { }

    for disc in discs:

        key = f"{disc.CoDisc} {disc.Abreviacao}"
        cod_mtr_sugestao[key] = []

        if disc.SemestreIdeal == semestre or \
                (semestre in [8, 7] and disc.TipoDisc == "optativaSI"):
            cods_tbl_hr[f"{disc.CoDisc} {disc.Abreviacao}"] = disc.CoDisc

        else:
            cods_tbl_hr_ext[f"{disc.CoDisc} {disc.Abreviacao}"] = disc.CoDisc
            mtr_auto_nome[disc.NomeDisc] = f"{disc.CoDisc} {disc.Abreviacao}"

    # Serve para listar as disciplinas na página
    if semestre in [7, 8]:
        vls_disciplinas = discs.filter(Q(SemestreIdeal=semestre) | Q(TipoDisc="optativaSI"))
    else:
        vls_disciplinas = discs.filter(SemestreIdeal=semestre)

    disc_obrig = vls_disciplinas.filter(TipoDisc="obrigatoria")
    tables_info = [
        {
            "title": "Disciplinas SI (Obrigatórias)",
            "mtrs": disc_obrig,
        },
        {
            "title": "Disciplinas SI (Optativas CB)",
            "mtrs": vls_disciplinas.filter(TipoDisc="optativaCB"),
        },
        {
            "title": "Disciplinas SI (Optativas SI)",
            "mtrs": vls_disciplinas.filter(TipoDisc="optativaSI"),
        },
    ]

    # prepara os dados para tabela de preferencias
    pref_semestre = Preferencias.objects.filter(
        CoDisc__SemestreIdeal=semestre
    ).order_by("CoDisc__Abreviacao")
    tbl_pref = [[disc.Abreviacao for disc in disc_obrig]]
    for row in range(1, 4):
        tbl_pref.append([])
        i = tbl_pref.index([])
        for _ in range(len(tbl_pref[0])):
            tbl_pref[i].append([])

        for pref in pref_semestre:
            if pref.nivel == row:
                j = tbl_pref[0].index(pref.CoDisc.Abreviacao)
                tbl_pref[row][j].append(pref.NumProf.Apelido)

    ano_func = AnoAberto.objects.get(id=1).Ano
    cod_mtr_sugestao = gera_sugestoes(ano_func, "sem_tds_profs")
    cod_mtr_sugestao_completo = gera_sugestoes(ano_func, "com_tds_profs")

    context = {
        "rest_horarios": restricoes_profs,
        "tbl_pref": tbl_pref,
        "semestre": semestre,
        "tProfs": tbl_vls,
        "auto_profs": auto_profs,
        "detalhes_profs": detalhes_profs,
        "cods_tbl_hr": cods_tbl_hr,
        "cods_tbl_hr_ext": cods_tbl_hr_ext,
        "mtr_auto_nome": mtr_auto_nome,
        "tables_info": tables_info,
        "impedimentos_totais": impedimentos_totais,
        "tbl_extra": tbl_extra,
        "anoAberto": AnoAberto.objects.get(id=1).Ano,
        "ano": ano,
        "cod_mtr_sugestao": cod_mtr_sugestao,
        "cod_mtr_sugestao_completo": cod_mtr_sugestao_completo
    }
    return render(request, "table/index.html", context)

def gera_sugestoes(ano, tds):
    discs = Disciplina.objects.all()
    cod_mtr_sugestao = {}

    for disc in discs:
        cod_mtr_sugestao[f"{disc.CoDisc} {disc.Abreviacao}"] = set()

    vls_turmas_auto = Turma.objects.filter(Ano=ano, CodTurma__in=[99, 98, 97])

    for vlr_auto in vls_turmas_auto:
        qtd_turma_manual = Turma.objects.filter(Ano=ano, NroUSP=vlr_auto.NroUSP, CoDisc=vlr_auto.CoDisc).exclude(
            CodTurma__in=[99, 98, 97]).count()
        qtd2_turma_auto = Turma.objects.filter(Ano=ano, NroUSP=vlr_auto.NroUSP, CoDisc=vlr_auto.CoDisc,
                                               CodTurma__in=(98, 99, 97)).count()
        if tds == "com_tds_profs":
            key = f"{vlr_auto.CoDisc.CoDisc} {vlr_auto.CoDisc.Abreviacao}"
            cod_mtr_sugestao[key].add(vlr_auto.NroUSP.Apelido)

        elif qtd_turma_manual < qtd2_turma_auto:
            key = f"{vlr_auto.CoDisc.CoDisc} {vlr_auto.CoDisc.Abreviacao}"
            cod_mtr_sugestao[key].add(vlr_auto.NroUSP.Apelido)

    #convertendo para lista os sets pois o json não aceita esse tipo
    for key, value in cod_mtr_sugestao.items():
        cod_mtr_sugestao[key] = list(value)

    return cod_mtr_sugestao


def atribuir_tbl_values(tbl, cod_disc, row, col, prof):
    tbl[row][
        col
    ] = f"{str(cod_disc.CoDisc)} {str(cod_disc.Abreviacao)}"
    tbl[row][col + 1] = prof

@login_required
def menu(request):
    ano = int(AnoAberto.objects.get(id=1).Ano)

    anos_ant = [i for i in range(2015, ano)]
    context = {
        "anos_ant": anos_ant,
        "anoAberto": ano,
        "sem_tur": turmas_obrigatórias_sem_horario(ano),
        "falta_aula": menos8_horas_aula_prof(ano)
    }
    return render(request, "table/menu.html", context)


def turmas_obrigatórias_sem_horario(ano):
    disciplinas = Disciplina.objects.filter(TipoDisc="obrigatoria", ativa=True).exclude(CoDisc__in=["ACH0021", "ACH0041"]).order_by('SemestreIdeal')
    dict_incompletas = {}
    for disc in disciplinas:
        turma02 = disc.turma_set.filter(Eextra="N", Ano=ano, CodTurma=2)
        turma04 = disc.turma_set.filter(Eextra="N", Ano=ano, CodTurma=4)
        turma94 = disc.turma_set.filter(Eextra="N", Ano=ano, CodTurma=94)

        result_string = ""

        if not turma02:
            result_string = result_string + "02"
            
        if not turma04:
            if result_string == "":
                result_string = result_string + "04"
            else:
                result_string = result_string + ", 04"


        if not turma94:
            if result_string == "":
                result_string = result_string + "94"
            else:
                result_string = result_string + ", 94"

        dict_incompletas[disc.Abreviacao] = {

            "disc": disc,
            "faltando": result_string,
            "smt": "impar" if disc.SemestreIdeal % 2 else "par"
        }

    return dict_incompletas


def menos8_horas_aula_prof(ano):
    #código readaptado do planilha_docentes.py
    profs = Professor.objects.filter(em_atividade=True)
    dicio = { }
    for prof in profs:

        pg_impar = prof.PG_1_semestre
        hr_rp1 = 4 * prof.rp1turmapreview_set.filter(ano=ano).count()
        hr_tadi = 2 * prof.taditurmapreview_set.filter(ano=ano).count()
        soma_impar = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="I").count() + pg_impar + hr_tadi + hr_rp1

        if soma_impar >= 8:
            soma_impar = -1

        pg_par = prof.PG_2_semestre
        hr_rp2 = 4 * prof.rp2turmapreview_set.filter(ano=ano).count()
        soma_par = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="P").count() + pg_par + hr_rp2

        if soma_par >= 8:
            soma_par = -1


        dicio[prof.Apelido] = {
            "I": soma_impar,
            "P": soma_par
        }
    return dicio


@login_required
def redirect(request):
    if request.method == "POST":
        valor_semestre = request.POST["select1"]
        ano = request.POST["anoSelecionado"]
        diretorio = str("./table/" + valor_semestre + "/" + ano + "/")
        return HttpResponseRedirect(diretorio)
    else:
        return HttpResponse("fail")

@login_required
def save_modify(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if not is_ajax:
        return HttpResponseBadRequest('Invalid request')
    if request.method != 'POST':
        return JsonResponse({'status': 'Invalid request'}, status=400)

    data = json.load(request)
    info_par = data["info"]

    ano = AnoAberto.objects.get(id=1).Ano

    erros = {}
    alertas = {}
    ind_modif = []

    if info_par["tipo"] == "d":
        if info_par["cod_disc"] == "ACH0042":
            deletar_valor_RP(data, ano, erros)
        else:
            deletar_valor(data, ano, erros)

    elif info_par["tipo"] == "i":


        aula_manha_noite(data, alertas, ano)
        aula_noite_outro_dia_manha(data, alertas, ano)

        if not aula_msm_horario(info_par, ano, data, erros):
        
            turma_obj = cadastrar_turma(info_par, ano, data["semestre"])
            update_prof(info_par, ano, data["semestre"])
            atualizar_dia(turma_obj, info_par, ano, erros, data["semestre"], ind_modif)

    elif info_par["tipo"] == "u":

        if "ant_prof" in info_par:
            aula_manha_noite(data, alertas, ano)
            aula_noite_outro_dia_manha(data, alertas, ano)

            if not aula_msm_horario(info_par, ano, data, erros):
               
                    turma_obj = update_prof(info_par, ano, data["semestre"])
                    indice_tbl_update(turma_obj, ind_modif, info_par)

        elif "ant_cod" in info_par:
            update_cod(data, ano, erros, data["semestre"], ind_modif)

    cod_mtr_sugestao = gera_sugestoes(ano, "sem_tds_profs")

    return JsonResponse({'erros': erros, 'alertas': alertas, 'cells_modif': ind_modif, 'cod_mtr_sugestao': cod_mtr_sugestao})

@login_required
def download_zip_planilhas(request):
    # content-type of response
    if request.method == "POST":
        ano = request.POST["ano_xlsx"]
        # cria o arquivo zip
        z = zipfile.ZipFile('Planilhas_graduação_SI.zip', 'w', zipfile.ZIP_DEFLATED)

        # planilha de docentes
        cwd = os.getcwd()
        file_path = os.path.join(cwd, "ferramentas/ferramenta_graduacao_si/table/static/table/docentes.xlsx")
        source_workbook = load_workbook(filename=file_path)
        sheet_doc = source_workbook["docentes"]
        planilha_docentes(sheet_doc)
        source_workbook.save("Docentes.xlsx")
        source_workbook.close()
        z.write("Docentes.xlsx")

        # planilha de distribuição semestres pares
        planilha_distribuição_semestre("par", ano)
        z.write("Distribuição_par.xlsx")

        # planilha de distribuição semestres impares
        planilha_distribuição_semestre("impar", ano)
        z.write("Distribuição_impar.xlsx")

        z.close()

        zip_arc = open("Planilhas_graduação_SI.zip", "rb")
        response = HttpResponse(zip_arc, content_type='application/x-gzip')
        response["Content-Disposition"] = 'attachment; filename="Planilhas_graduação_SI.zip"'

        # limpa os arquivos criados no processo
        os.remove("Docentes.xlsx")
        os.remove("Distribuição_impar.xlsx")
        os.remove("Distribuição_par.xlsx")
        os.remove("Planilhas_graduação_SI.zip")

        return response
    else:
        return "erro"


def planilha_distribuição_semestre(semestre, ano):
    wb = openpyxl.Workbook()
    sheet_si = wb.active
    sheet_si.title = "SI"
    planilha_si(sheet_si, semestre, ano)
    sheet_extra = wb.create_sheet("Extra")
    planilha_extra(sheet_extra, ano)
    wb.save(f"Distribuição_{semestre}.xlsx")
    wb.close()


@login_required
def pref_planilha(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file", None)
        excel_type = request.POST["excel_type"]

        if not excel_file:
            return render(
                request, "table/menu.html", {"error_message": "Nenhum arquivo enviado."}
            )

        if not excel_file.name.endswith(".xlsx"):
            return render(
                request,
                "table/menu.html",
                {
                    "error_message": "Formato de arquivo inválido. Por favor, envie um arquivo do tipo .xlsx."
                },
            )

        try:
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            profs = Professor.objects.all()
            ano = AnoAberto.objects.get(id=1).Ano

            header = [cell.value for cell in worksheet[1]]
            if excel_type == "pref_disc_hro":
                Preferencias.objects.filter(AnoProf=ano).delete()
                Restricao.objects.all().delete()
                MtvRestricao.objects.all().delete()

            for row in worksheet.iter_rows(min_row=2, max_col=39, values_only=True):

                email = row[1]
                if not isinstance(email, str) or email == "":
                    continue

                prof_encontrado = False
                email_professor = False

                # consultar direto o email
                for prof in profs:
                    if email == prof.Email:
                        email_professor = email
                        prof_encontrado = True

                if not prof_encontrado:
                    print(f"PROFESSOR EMAIL { email_professor } NÃO ENCONTRADO")
                    continue

                # para o semestre_par sendo false a planilha mudou para uma nova
                # o código abaixo que pega os dados da planilha antiga estão comentados
                # caso queira utilizar ela para carregar restrições, por exemplo, mude os comentários

                semestre_par = True if excel_type == "pref_hro_2" else False

                prof_db = Professor.objects.get(Email=email_professor)

                if semestre_par:
                    prof_db.consideracao2 = row[9]
                else:
                    # prof_db.consideracao1 = row[37]
                    # prof_db.pos_doc = row[33]
                    # prof_db.pref_optativas = row[32]

                    prof_db.consideracao1 = row[78]
                    if row[31]:
                        justificativaMenos8Horas(professor=prof_db, ano=ano, semestre_ano="P", texto_justificando=row[33]).save()
                    elif row[32]:
                        justificativaMenos8Horas(professor=prof_db, ano=ano, semestre_ano="I", texto_justificando=row[33]).save()

                    if row[34]:
                        justificativaMenos8Horas(professor=prof_db, ano=ano, semestre_ano="P",
                                                 texto_justificando=row[36]).save()
                    elif row[35]:
                        justificativaMenos8Horas(professor=prof_db, ano=ano, semestre_ano="I",
                                                 texto_justificando=row[36]).save()

                    pref_disc_excel_impar("impar", row, prof_db, header)
                    pref_disc_excel_impar("par", row, prof_db, header)

                prof_db.save()
                pref_horarios(row, prof_db, semestre_par)

                # print(f"coluna AI{row[34]}, coluna AK{pref_horario}")
            return render(
                request,
                "table/menu.html",
                {"success_message": "Dados de preferência salvos com sucesso."},
            )
        except Exception as e:
            print(e)
            return render(
                request,
                "table/menu.html",
                {"error_message": "Ocorreu um erro ao processar o arquivo."},
            )

    return render(request, "table/menu.html")


