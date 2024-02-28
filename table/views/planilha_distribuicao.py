from table.models import *
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment


def planilha_extra(sheet, ano):
    # Filtra as turmas do semestre desejado, trazendo os dados necessários e ordenando-os.
    turmas_queryset = (
        Turma.objects.filter(
            Ano=ano,
            Eextra="S",
        )
        .prefetch_related("dia_set", "CoDisc")
        .order_by("CoDisc__SemestreIdeal", "CodTurma")
    )
    row = create_header(sheet, 1, "TURMAS EXTRAS")
    largura_cols(sheet)
    escreve_excel(turmas_queryset, sheet, True, row)


def planilha_si(sheet_si, smt, ano):
    semestre_geral = [1, 3, 5, 7] if smt == "impar" else [2, 4, 6, 8]

    # Filtra as turmas do semestre desejado, trazendo os dados necessários e ordenando-os.
    turmas_queryset = (
        Turma.objects.filter(
            Ano=ano,
            CoDisc__SemestreIdeal__in=semestre_geral,
            Eextra="N",
        )
        .prefetch_related("dia_set", "CoDisc")
        .order_by("CoDisc__SemestreIdeal", "CoDisc__CoDisc", "CodTurma")
    )
    print(turmas_queryset)

    format_border(sheet_si, "A1")
    sheet_si.merge_cells("A1:E2")
    sheet_si["A1"].font = Font(size=15, name="Arial", bold=True)
    sheet_si["A1"].alignment = Alignment("center", "center", wrap_text=True)
    sheet_si[
        "A1"
    ].value = f"DISTRIBUIÇÃO DE CARGA HORÁRIA  – {semestre_geral[0]}º/Sem/{datetime.now().year}"

    format_border(sheet_si, "A3")
    sheet_si.merge_cells("A3:E3")
    sheet_si["A3"].value = "DISCIPLINAS ESPECÍFICAS - SISTEMAS DE INFORMAÇÃO"
    sheet_si["A3"].font = Font(size=13, name="Arial", bold=True)
    sheet_si["A3"].fill = PatternFill("solid", fgColor="00C0C0C0")
    sheet_si["A3"].alignment = Alignment("center", "center", wrap_text=True)
    sheet_si.row_dimensions[3].height = 20

    largura_cols(sheet_si)
    escreve_excel(turmas_queryset, sheet_si, False, 4)


def format_border(self, coord):
    border = Border(
        left=Side(border_style="thin", color="00000000"),
        right=Side(border_style="thin", color="00000000"),
        top=Side(border_style="thin", color="00000000"),
        bottom=Side(border_style="thin", color="00000000"),
    )
    self[coord].border = border


def largura_cols(sheet):
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 13
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["D"].width = 24
    sheet.column_dimensions["E"].width = 22


def insere_linha_branca(self, row):
    format_border(self, "A" + str(row))
    self.merge_cells("A" + str(row) + ":E" + str(row))
    self.row_dimensions[row].height = 12
    return row + 1


def create_header(sheet_si, row, title):
    coord = "A" + str(row)
    format_border(sheet_si, coord)
    sheet_si.merge_cells(coord + ":E" + str(row))
    sheet_si[coord].value = title
    sheet_si[coord].font = Font(size=13, name="Arial", bold=True)
    sheet_si[coord].fill = PatternFill("solid", fgColor="00C0C0C0")
    sheet_si[coord].alignment = Alignment("center", "center", wrap_text=True)
    sheet_si.row_dimensions[row].height = 20

    row += 1
    sheet_si.row_dimensions[row].height = 13
    l_header = ["DISCIPLINA", "TURMA", "HORÁRIO", "DOCENTES", "PRÉDIOS / SALAS"]
    cols = ["A", "B", "C", "D", "E"]

    for letter, header in zip(cols, l_header):
        coord = letter + str(row)
        sheet_si[coord].value = header
        sheet_si[coord].font = Font(size=9, name="Arial", bold=True)
        sheet_si[coord].alignment = Alignment("center", "center", wrap_text=True)
        format_border(sheet_si, coord)

    row += 1
    return row


def conct_horario(dias):
    if dias.count() > 1 and dias[0].DiaSemana == dias[1].DiaSemana:
        hro_inicial = dias[0].get_Horario_display().split("-")[0]
        hro_final = dias[1].get_Horario_display().split("-")[1]
        hro = f"{dias[0].get_DiaSemana_display()[0:3]} {hro_inicial} - {hro_final}"
        dias = list(dias)
        del dias[0]
        return hro
    else:
        return ""


def cria_horarios(hro_msm_dia, dias):
    horarios = []

    if hro_msm_dia:
        horarios.append(hro_msm_dia)
    else:
        horarios = [
            f"{dia.get_DiaSemana_display()[0:3]} {dia.get_Horario_display()}"
            for dia in dias
        ]
    return horarios


def format_cell(self, coord):
    alignment = Alignment("center", "center", wrap_text=True)
    format_border(self, coord)
    self[coord].alignment = alignment
    self[coord].font = Font(size=9, name="Arial")


def set_col_a(sheet_si, disc, ini, end):
    # formata a coordenada da coluna A para o nome da matéria
    coord = "A" + str(ini)
    format_cell(sheet_si, coord)
    if ini < end:
        merge_coor = "A" + str(ini) + ":" + "A" + str(end - 1)
        sheet_si.merge_cells(merge_coor)

    sheet_si[coord].value = disc.CoDisc + " - " + disc.NomeDisc


def preenche_celulas(sheet_si, qtd_linhas, row, turma):
    cols = {"B": turma.CodTurma, "D": turma.NroUSP.Apelido, "E": ""}
    for col, value in cols.items():
        coord = col + str(row) + ":" + col + str(row + qtd_linhas - 1)
        cell = col + str(row)
        format_cell(sheet_si, cell)
        sheet_si.merge_cells(coord)
        sheet_si[cell].value = value


def escreve_excel(turmas_queryset, sheet_si, extra, row):
    # Variáveis para controle de semestre e início da tabela da disciplina atual.
    smt_atual = -1
    row_ini = row if extra else -1

    # Iteração pelas turmas, criando as tabelas de horários das disciplinas.
    for i, turma in enumerate(turmas_queryset):
        if not turma.dia_set.all():
            print(f"turma sem nenhum dia: {turma}")
            continue
        # Obtém a disciplina da turma atual.
        disc = turma.CoDisc

        # Se o semestre da disciplina atual for diferente do anterior,
        # insere uma linha em branco e o título do semestre atual.
        if smt_atual != disc.SemestreIdeal and not extra:
            row = insere_linha_branca(sheet_si, row)
            title = f"{disc.SemestreIdeal}° SEMESTRE"
            row = create_header(sheet_si, row, title)
            smt_atual = disc.SemestreIdeal
            row_ini = row

        # Obtém os dias de aula da turma atual, concatenando os horários de aula nos dias em que ocorrem no mesmo dia.
        dias = turma.dia_set.all().order_by("Horario")
        hro_msm_dia = conct_horario(dias)

        # Cria uma lista com os horários de aula da turma atual.
        horarios_turma = cria_horarios(hro_msm_dia, dias)

        # Iteração pelos horários de aula da turma atual,
        # preenchendo a tabela de horários com os valores da disciplina atual.
        for horario in horarios_turma:
            row_copy = row
            sheet_si.row_dimensions[row_copy].height = 30
            coord = "C" + str(row_copy)
            format_cell(sheet_si, coord)
            sheet_si[coord].value = horario
            row += 1

        # Preenche as células da tabela de horários com as informações da disciplina atual.
        preenche_celulas(
            sheet_si, len(horarios_turma), row - len(horarios_turma), turma
        )

        # Verifica se existe uma próxima turma, e se for o caso, se é uma turma de outra disciplina.
        # Caso afirmativo, configura a coluna A da disciplina.
        prox_turma = turmas_queryset[i + 1] if i + 1 < len(turmas_queryset) else None
        if prox_turma is not None and str(prox_turma.CoDisc) != str(disc):
            set_col_a(sheet_si, disc, row_ini, row)
            row_ini = row
        elif prox_turma is None:
            set_col_a(sheet_si, disc, row_ini, row)

    # Define a área de impressão da tabela.
    if turmas_queryset:
        sheet_si.print_area = f"A1:E{row - 1}"
