#FUNÇÕES PARA ATRIBUIÇÃO AUTOMÁTICA
import os
import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.db import IntegrityError
from django.db.models.functions import Coalesce
from django.http import HttpResponseRedirect
from django.shortcuts import render
from openpyxl.reader.excel import load_workbook
from unidecode import unidecode
import math
from ..models import *
from django.db.models import Prefetch, Count, Q, Max, Value
from datetime import datetime
from openpyxl.styles import PatternFill, Font


def planilha_docentes(sheet_doc):
    # carregando os dados na memória
    ano_funcionamento = Ano=AnoAberto.objects.get(id=1).Ano
    turma_prefetch = Prefetch(
        "turma_set",
        queryset=Turma.objects.filter(Ano=ano_funcionamento).select_related("CoDisc"),
    )
    profs = (
        Professor.objects.filter(em_atividade=True).prefetch_related(turma_prefetch).order_by("NomeProf")
    )

    row = 3
    for professor in profs:
        col_1_disc = 2
        col_2_disc = 8
        hrs_aula_1 = 0
        hrs_aula_2 = 0


        #Precisa cuidar do caso do optativaCB
        for turma in professor.turma_set.all():
            if turma.CoDisc.TipoDisc == "optativaCB":
                continue

            if turma.CoDisc.SemestreIdeal % 2:
                # Impar
                hrs_aula_1 += preenche_doc(sheet_doc, row, col_1_disc, turma)
                col_1_disc += 1
            else:
                hrs_aula_2 += preenche_doc(sheet_doc, row, col_2_disc, turma)
                col_2_disc += 1

        for turma_rp1 in professor.rp1turmapreview_set.all():
            hrs_aula_1 += preenche_doc_rp1(sheet_doc, row, col_1_disc, turma_rp1)
            col_1_disc += 1

        for turma_rp1 in professor.rp2turmapreview_set.all():
            sheet_doc.cell(
                row=row, column=col_2_disc
            ).value = f"ACH0042-RP2"
            hrs_aula_2 += 4
            col_2_disc += 1

        for turma_tadi in professor.taditurmapreview_set.all():
            sheet_doc.cell(
                row=row, column=col_1_disc
            ).value = f"ACH0021-TADI"
            hrs_aula_1 += 2
            col_1_disc += 1

        sheet_doc.cell(row=row, column=1).value = professor.NomeProf

        #Poderia ser qualquer uma, só serve para indicar o semestre
        ia = Disciplina.objects.get(Abreviacao="IA")
        mqa = Disciplina.objects.get(Abreviacao="MQA")

        if not mais8_horas_aula_prof(ia, professor):
            sheet_doc.cell(row=row, column=col_1_disc).font = Font(
                color="FF0000", size=10, name="Calibri"
            )
            sheet_doc.cell(
                row=row, column=col_1_disc
            ).value = "Extra(s)/Optativa(s)"

        if not mais8_horas_aula_prof(mqa, professor):

            sheet_doc.cell(row=row, column=col_2_disc).font = Font(
                color="FF0000", size=10, name="Calibri"
            )
            sheet_doc.cell(
                row=row, column=col_2_disc
            ).value = "Extra(s)/Optativa(s)"

        justificativas = justificativaMenos8Horas.objects.filter(professor=professor, ano=ano_funcionamento)
        for justif in justificativas:

            if justif.semestre_ano == "P":
                preenche_justificativa(sheet_doc, professor, justif, row, col_2_disc, 4)

            if justif.semestre_ano == "I":
                preenche_justificativa(sheet_doc, professor, justif, row, col_1_disc, 5)


        # Preenche os créditos aula
        sheet_doc[f"V{row}"].value = hrs_aula_1 + professor.PG_1_semestre
        sheet_doc[f"W{row}"].value = hrs_aula_2 + professor.PG_2_semestre
        sheet_doc[f"X{row}"].value = hrs_aula_2 + hrs_aula_1 + professor.PG_1_semestre + professor.PG_2_semestre
        sheet_doc[f"G{row}"].value = professor.PG_1_semestre if professor.PG_1_semestre != 0 else ""
        sheet_doc[f"L{row}"].value = professor.PG_2_semestre if professor.PG_2_semestre != 0 else ""

        row += 1


def preenche_justificativa(pag, p, justificativa, row, col, v):
    pag.cell(
        row=row, column=col
    ).value = f"{str(justificativa.get_justificativa_display())}"

    fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")

    for i in range(0, v):
        pag.cell(row=row, column=col + i).fill = fill


def preenche_doc(sheet_doc, row, col, turma):
    disciplina = turma.CoDisc
    if turma.Eextra == "S":
        sheet_doc.cell(row=row, column=col).font = Font(
            color="FF0000", size=10, name="Calibri"
        )

    if disciplina.TipoDisc.lower() != "obrigatoria":
        sheet_doc.cell(row=row, column=col).fill = openpyxl.styles.PatternFill(
            start_color="008000", end_color="008000", fill_type="solid"
        )
        sheet_doc.cell(row=row, column=col).font = Font(color="000000", underline="none")

    sheet_doc.cell(
        row=row, column=col
    ).value = f"{str(disciplina.CoDisc)}-{str(disciplina.Abreviacao)}"

    return disciplina.CreditosAula


def preenche_doc_rp1(sheet_doc, row, col, turma):
    sheet_doc.cell(
        row=row, column=col
    ).value = f"ACH0041-RP1"

    return 4

def criar_atribuicoes(request):
    discs = Disciplina.objects.filter(ativa=True, TipoDisc="obrigatoria")
    ano_atual = int(AnoAberto.objects.get(id=1).Ano)

    Turma.objects.filter(CodTurma__in=(99, 98, 97), Ano=ano_atual).delete()
    
    TadiTurmaPreview.objects.filter(codigo__in=(range(1, 17)), ano=ano_atual).delete()
    RP1TurmaPreview.objects.filter(codigo=99, ano=ano_atual).delete()
    nova_turma = RP1TurmaPreview.objects.create(
        codigo=99,
        ano=ano_atual
    )

    n_tadi = n_rp1 = 1

    #Vai tentar completar com os i anos anteriores os profs sem preferências
    for i in range(4):
        n_tadi, n_rp1 = cria_atribuicao_profs_sem_pref(ano_atual - i, n_tadi, n_rp1, nova_turma)

    cria_atribuicoes_obrigatorias(ano_atual)

    # As duas funçoes abaixo substituem a imediatamente acima
    # Elas são mais pesadas e tem desempenho pior, para rodar precisa descomentar
    # os códigos do fim desse arquivo
    # cria_atribuicao_com_pref(discs, ano_atual)
    completa_atrib_obrig_com_hist(discs, ano_atual)

    cria_atribuicao_com_pref_rp1(nova_turma)
    cria_atribuicao_com_pref_rp2()

    cria_atribuicao_com_pref_tadi(n_tadi)
    completa_atrib_tadi_com_hist(ano_atual)

    cria_atribuicoes_optativas(ano_atual)

    print("execução da atribuição automática finalizada")
    return HttpResponseRedirect("/#item-0")

def calcular_pontuacao(preferencia, vezes_dada, max_vezes):
    # Pesos para a fórmula
    w1 = 0.5
    w2 = 0.5
    pontuacao = w1 / preferencia + w2*(vezes_dada / max_vezes)
    if pontuacao <= 0.01:
        pontuacao = 0

    return pontuacao


def cria_atribuicoes_optativas(ano_atual):
    # Obtém as disciplinas optativas do tipo 'optativaSI' que estão ativas
    disciplinas_opt = Disciplina.objects.filter(TipoDisc='optativaSI', ativa=True)

    for disciplina in disciplinas_opt:
        # Obtém todas as preferências dos professores para a disciplina atual
        preferencias = Preferencias.objects.filter(CoDisc=disciplina, AnoProf=ano_atual)
        semestre = "P" if disciplina.SemestreIdeal % 2 == 0 else "I"

        # Encontra o primeiro professor com preferência e menos de 8 horas de aula
        for pref in preferencias:

            professor = pref.NumProf
            p_justf = justificativaMenos8Horas.objects.filter(semestre_ano=semestre, ano=ano_atual, professor=pref.NumProf)

            if not mais8_horas_aula_prof(disciplina, professor) and not p_justf:

                # Cria uma turma para a disciplina com o professor selecionado
                Turma.objects.create(
                    CoDisc=disciplina,
                    CodTurma=99,
                    Ano=ano_atual,
                    NroUSP=professor,
                    Eextra="N",
                    semestre_extra=disciplina.SemestreIdeal,
                    SemestreAno=semestre
                )


def cria_atribuicoes_obrigatorias(ano_funcionamento):

    # Encontrar todas as disciplinas obrigatórias
    disciplinas_obrigatorias = (Disciplina.objects.filter(TipoDisc='obrigatoria', ativa=True)
                                .exclude(CoDisc__in=("ACH0041", "ACH0021", "ACH0042")))

    # Encontrar o número máximo de vezes que qualquer professor já deu qualquer matéria
    max_vezes = (
        Turma.objects.values('CoDisc', 'NroUSP')
        .annotate(total_vezes=Count('NroUSP'))
        .aggregate(max_vezes=Coalesce(Max('total_vezes'), Value(1)))['max_vezes']
    )

    for disciplina in disciplinas_obrigatorias:

        turmas_disc = Turma.objects.filter(CoDisc=disciplina, Ano=ano_funcionamento, Eextra="N").count()
        if turmas_disc >= 3:
            continue

        smt = "P" if disciplina.SemestreIdeal % 2 == 0 else "I"
        profs_list = profs_mais_8hrs(disciplina)
        todos_professores = Professor.objects.filter(em_atividade=True).exclude(NomeProf__in=profs_list)

        pontuacoes = []
        for professor in todos_professores:
            if justificativaMenos8Horas.objects.filter(professor=professor, ano=ano_funcionamento, semestre_ano=smt):
                continue

            preferencia = Preferencias.objects.filter(CoDisc=disciplina, NumProf=professor).first()
            if preferencia:
                nivel = preferencia.nivel
            else:
                nivel = 999999

            # Calcular o número de vezes que o professor já deu a matéria
            vezes_dada = Turma.objects.filter(CoDisc=disciplina, NroUSP=professor).count()

            pontuacao = calcular_pontuacao(nivel, vezes_dada, max_vezes)
            if pontuacao:
                pontuacoes.append((professor, pontuacao))

        # Seleciona os três professores com a maior pontuação
        pontuacoes.sort(key=lambda x: x[1], reverse=True)
        print(pontuacoes)
        if len(pontuacoes) >= 3:
            top_professores = [prof[0] for prof in pontuacoes[:3]]
        else:
            top_professores =  [prof[0] for prof in pontuacoes]

        print(f"{disciplina.Abreviacao} - {top_professores}")

        # Cria três turmas para a disciplina obrigatória
        t_faltando = 3 - turmas_disc

        for i in range(t_faltando):
            try:
                n = Turma.objects.create(
                    CoDisc=disciplina,
                    CodTurma=98,
                    Ano=ano_funcionamento,
                    NroUSP=top_professores[i],
                    NroAlunos=30,
                    Eextra='N',
                    semestre_extra=disciplina.SemestreIdeal,
                    SemestreAno=smt
                )
            except:
                pass


def cria_atribuicao_com_pref_rp1(nova_turma):
    ano_atual = AnoAberto.objects.get(id=1).Ano
    rp1 = Disciplina.objects.get(CoDisc="ACH0041")
    p_justf = justificativaMenos8Horas.objects.filter(semestre_ano="I").values_list('professor', flat=True).distinct()

    for prioridade in range(1, 4):

        #Se a ordem mudar a função abaixo deve ser implementada
        profs_list = profs_mais_8hrs(rp1)
        pref_disc = (rp1.preferencias_set.filter(AnoProf=ano_atual, nivel=prioridade)
                     .exclude(Q(NumProf__in=profs_list) | Q(NumProf__in=p_justf)))

        num_pref = pref_disc.count()

        #considerando que só haja uma instância q fica gravada por ano
        num_profs = nova_turma.professor_si.count()

        if num_profs >= 6:
            continue

        profs_faltando = 6 - num_profs

        if num_pref > profs_faltando:

            hist = Professor.objects.annotate(num_turmas=Count('rp1turma')).order_by('-num_turmas')

            for i in range(0, profs_faltando):
                nova_turma.professor_si.add(hist[i])

        else:
            for i in range(0, num_pref):
                nova_turma.professor_si.add(pref_disc[i].NumProf)


def cria_atribuicao_com_pref_rp2():

    ano_atual = AnoAberto.objects.get(id=1).Ano
    rp2 = Disciplina.objects.get(CoDisc="ACH0042")
    RP2TurmaPreview.objects.filter(codigo=99, ano=ano_atual).delete()
    nova_turma = RP2TurmaPreview.objects.create(
        # Código arbitrário
        codigo=99,
        ano=ano_atual
    )

    p_justf = justificativaMenos8Horas.objects.filter(semestre_ano="P").values_list('professor', flat=True).distinct()

    for prioridade in range(1, 4):

        #Se a ordem mudar a função abaixo deve ser implementada
        profs_list = profs_mais_8hrs(rp2)


        pref_disc = (rp2.preferencias_set.filter(AnoProf=ano_atual, nivel=prioridade)
                     .exclude(Q(NumProf__in=profs_list) | Q(NumProf__in=p_justf)))

        num_pref = pref_disc.count()
        #considerando que só haja uma instância q fica gravada por ano
        num_profs = nova_turma.professor_si.count()

        if num_profs >= 6:
            continue

        profs_faltando = 6 - num_profs

        for pref in pref_disc:
            nova_turma.professor_si.add(pref.NumProf)


def mais8_horas_aula_prof(disc, prof):
    mais_8hrs = []
    ano = AnoAberto.objects.get(id=1).Ano
    smt = "I" if disc.SemestreIdeal in (1, 3, 5, 7) else "P"

    if smt == "I":
        # se a pg for 2h vale desconsiderá-la para o cálculo
        pg_impar = 0 if prof.PG_1_semestre == 2 else prof.PG_1_semestre
        hr_rp1 = 4 * prof.rp1turmapreview_set.filter(ano=ano).count()
        hr_tadi = 2 * prof.taditurmapreview_set.filter(ano=ano).count()
        prof_hrs = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="I").count() + pg_impar + hr_tadi + hr_rp1

    else:
        pg_par = 0 if prof.PG_2_semestre == 2 else prof.PG_2_semestre
        hr_rp2 = 4 * prof.rp2turmapreview_set.filter(ano=ano).count()
        prof_hrs = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="P").count() + pg_par + hr_rp2

    if prof_hrs >= 8:
        return True


def profs_mais_8hrs(disc):
    mais_8hrs = []
    ano = AnoAberto.objects.get(id=1).Ano

    profs = Professor.objects.all()
    smt = "I" if disc.SemestreIdeal in (1, 3, 5, 7) else "P"

    for prof in profs:
        if smt == "I":
            # se a pg for 2h vale desconsiderá-la para o cálculo
            pg_impar = 0 if prof.PG_1_semestre == 2 else prof.PG_1_semestre
            hr_rp1 = 4 * prof.rp1turmapreview_set.filter(ano=ano).count()
            hr_tadi = 2 * prof.taditurmapreview_set.filter(ano=ano).count()
            prof_hrs = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="I").count() + pg_impar + hr_tadi + hr_rp1

        else:
            pg_par = 0 if prof.PG_2_semestre == 2 else prof.PG_2_semestre
            hr_rp2 = 4 * prof.rp2turmapreview_set.filter(ano=ano).count()
            prof_hrs = 4 * prof.turma_set.filter(Ano=ano, SemestreAno="P").count() + pg_par + hr_rp2

        if prof_hrs >= 8:
            mais_8hrs.append(prof)

    return mais_8hrs


def cria_atribuicao_com_pref_tadi(num_turma):

    ano_atual = AnoAberto.objects.get(id=1).Ano
    tadi = Disciplina.objects.get(CoDisc="ACH0021")

    p_justf = justificativaMenos8Horas.objects.filter(semestre_ano="I").values_list('professor', flat=True).distinct()

    for prioridade in range(1, 4):

        profs_list = profs_mais_8hrs(tadi)

        pref_disc = (tadi.preferencias_set.filter(AnoProf=ano_atual, nivel=prioridade)
                     .exclude(Q(NumProf__in=profs_list) | Q(NumProf__in=p_justf)))

        num_pref = pref_disc.count()

        turmas_tadi = TadiTurmaPreview.objects.filter(ano=ano_atual).count()

        if turmas_tadi >= 16:
            continue

        turmas_faltando = 16 - turmas_tadi

        if num_pref > turmas_faltando:

            for i in range(0, turmas_faltando):

                hist = Professor.objects.annotate(num_turmas=Count('taditurma')).order_by('-num_turmas')

                y = num_turma + 2
                while num_turma < y:

                    nova_turma = TadiTurmaPreview.objects.create(
                        codigo=num_turma,
                        ano=ano
                    )

                    nova_turma.professor_si.add(hist[i].NumProf)
                    num_turma = num_turma + 1
        else:
            for i in range(0, num_pref):
                cadastra_turmas_auto_tadi(pref_disc[i], num_turma)
                num_turma += 2


def cadastra_turmas_auto_tadi(pref, num):
    ano = AnoAberto.objects.get(id=1).Ano

    y = num + 2
    while num < y:
        nova_turma = TadiTurmaPreview.objects.create(
            codigo=num,
            ano=ano
        )

        nova_turma.professor_si.add(pref.NumProf)
        num += 1


def completa_atrib_tadi_com_hist(ano):

    disc = Disciplina.objects.get(CoDisc="ACH0021")
    num_turmas = TadiTurmaPreview.objects.filter(ano=ano).count()

    if num_turmas >= 16:
        return

    profs_list = profs_mais_8hrs(disc)
    p_justf = justificativaMenos8Horas.objects.filter(semestre_ano="I").values_list('professor', flat=True).distinct()

    hist = Professor.objects.exclude(Q(NomeProf__in=profs_list) | Q(p_justf)).annotate(num_turmas=Count('taditurma')).order_by('-num_turmas')
    t_faltando = math.floor((16 - num_turmas) / 2)

    i = 0
    cod_turma = num_turmas + 1

    while i < t_faltando:
        try:
            y = cod_turma + 2
            while cod_turma < y:
                nova_turma = TadiTurmaPreview.objects.create(
                    codigo=cod_turma,
                    ano=ano
                )

                nova_turma.professor_si.add(hist[i])
                cod_turma = cod_turma + 1

            i = i + 1
        except IntegrityError:

            if t_faltando + 1 <= hist.count():
                t_faltando = t_faltando + 1
            else:
                print(f"Disciplina {disc.Abreviacao} faltando turmas")

            i = i + 1


def cria_atribuicao_profs_sem_pref(ano_atual, num_tadi, num_rp1, rp1_turma):
    profs_sem_pref = Professor.objects.filter(em_atividade=True).exclude(
        Q(preferencias__isnull=False)
    ).distinct()

    for prof in profs_sem_pref:
        p_turmas = (prof.turma_set.filter(Ano=ano_atual - 1, CoDisc__TipoDisc="obrigatoria")
                    .exclude(CoDisc__in=("ACH0041", "ACH0021", "ACH0042")))

        for t in p_turmas:
            turmas_disc = Turma.objects.filter(CoDisc=t.CoDisc, Ano=ano_atual, Eextra="N").count()

            if not mais8_horas_aula_prof(t.CoDisc, prof) and turmas_disc < 3:
                cadastra_turma_ano_anterior(t)

        for _ in prof.rp1turma_set.filter(ano=ano_atual - 1):
            if num_rp1 < 6:
                rp1_turma.professor_si.add(prof)
                num_rp1 += 1

        for _ in prof.taditurma_set.filter(ano=ano_atual - 1):
            if num_tadi < 16:
                nv_turma = TadiTurmaPreview.objects.create(
                    ano=ano_atual,
                    codigo=num_tadi
                )
                nv_turma.professor_si.add(prof)
                num_tadi += 1

        # RP2
        # for t in prof.rp1turma_set.filter(Ano=ano_atual - 1):
        #     if hrs_impar < 8:
        #         RP1TurmaPreview.objects.create(
        #             professor_si = prof,
        #             ano = ano_atual
        #         )
        #         hrs_impar += 4

    return [num_tadi, num_rp1]


def cadastra_turma_ano_anterior(t):
    ano_atual = AnoAberto.objects.get(id=1).Ano
    try:
        nova_turma = Turma.objects.create(
            CoDisc=t.CoDisc,
            # Código arbitrário
            CodTurma=99,
            Ano=ano_atual,
            NroUSP=t.NroUSP,
            Eextra="N",
            SemestreAno=t.SemestreAno,
            semestre_extra=t.CoDisc.SemestreIdeal
        )

    except IntegrityError as e:
        try:
            nova_turma = Turma.objects.create(
                CoDisc=t.CoDisc,
                # Código arbitrário
                #para turmas iguais com diferença do código por conta do horário
                CodTurma=98,
                Ano=ano_atual,
                NroUSP=t.NroUSP,
                Eextra="N",
                SemestreAno=t.SemestreAno,
                semestre_extra=t.CoDisc.SemestreIdeal
            )
        except:
            pass
    finally:
        pass


@login_required
def carregar_atribuicao(request):
    if not request.method == "POST":
        return render(request, "table/menu.html")

    excel_file = request.FILES.get("excel_file", None)
    if not excel_file:
        return render(
            request, "table/menu.html", {"error_message": "Nenhum arquivo enviado."}
        )

    if not excel_file.name.endswith(".xlsx"):
        return render(
            request,
            "table/menu.html",
            {
                "error_message": "Formato de arquivo inválido. Por favor, envie um arquivo do tipo .xlsx."
            },
        )

    source_workbook = load_workbook(filename=excel_file)
    sheet_doc = source_workbook["docentes"]
    profs = Professor.objects.all()
    ano_atual = AnoAberto.objects.get(id=1).Ano
    Turma.objects.filter(CodTurma__in=(99, 98, 97), Ano=ano_atual).delete()
    TadiTurmaPreview.objects.filter(codigo__in=(range(1, 17)), ano=ano_atual).delete()
    RP1TurmaPreview.objects.filter(codigo=99, ano=ano_atual).delete()
    RP2TurmaPreview.objects.filter(codigo=99, ano=ano_atual).delete()

    num_turma = 1
    nova_turma = RP1TurmaPreview.objects.create(
        codigo=99,
        ano=ano_atual
    )

    rp2_turma = RP2TurmaPreview.objects.create(
        # Código arbitrário
        codigo=99,
        ano=ano_atual
    )


    l_justificativa = ("Extra(s)/Optativa(s)", None, "Afastado", "Compensação de Créditos", "Empréstimo", "Licença-Maternidade","Licença-Prêmio", "Sem Contrato", "Pós-Doutorado")

    for row in sheet_doc.iter_rows(min_row=3, max_row=40, max_col=11, values_only=True):
        if not row[0]:
            continue

        prof_obj = Professor.objects.get(NomeProf=row[0])

        # cria disciplina para o professor
        for disciplina in row[1:]:
            codisc_db = ""

            if disciplina is None or not isinstance(disciplina, str):
                continue

            try:
                codisc = disciplina.split("-")[0]
                codisc_db = Disciplina.objects.get(CoDisc=codisc)

            except ObjectDoesNotExist:
                # print(f"disciplina {disciplina} não encontrada para o professor: {prof_obj.Apelido}")
                continue


            #RP1
            if codisc == "ACH0041":
                nova_turma.professor_si.add(prof_obj)

            elif codisc == "ACH0021":
                tadi_turma = TadiTurmaPreview.objects.create(
                    codigo=num_turma,
                    ano=ano_atual
                )

                tadi_turma.professor_si.add(prof_obj)

                num_turma += 1

            elif codisc == "ACH0042":
                rp2_turma.professor_si.add(prof_obj)

            elif not codisc in l_justificativa:
                semestre = "P" if codisc_db.SemestreIdeal % 2 == 0 else "I"

                try:
                    n = Turma.objects.create(
                        CoDisc=codisc_db,
                        CodTurma=99,
                        Ano=ano_atual,
                        NroUSP=prof_obj,
                        NroAlunos=30,
                        Eextra='N', #precisa implementar para o sistema reconhecer como extra
                        semestre_extra=codisc_db.SemestreIdeal,
                        SemestreAno=semestre
                    )
                except IntegrityError:
                    try:
                        n = Turma.objects.create(
                            CoDisc=codisc_db,
                            CodTurma=98,
                            Ano=ano_atual,
                            NroUSP=prof_obj,
                            NroAlunos=30,
                            Eextra='N',  # precisa implementar para o sistema reconhecer como extra
                            semestre_extra=codisc_db.SemestreIdeal,
                            SemestreAno=semestre
                        )
                    except:
                        print(f"{codisc_db}\n NÃO FOI GRAVADA NA 3 vez")
                        pass

            # print("\n")

    verifica_exclusao(ano_atual)

    return HttpResponseRedirect("/#item-0")


def verifica_exclusao(ano):
    # Trata o caso de quando a modificação da planilha docentes.xlsx excluiu
    # alguma turma já criada nas grades

    #Grade geral
    turmas_gerais = Turma.objects.filter(Ano=ano).exclude(CoDisc__CoDisc__in=("ACH0041", "ACH0021", "ACH0042"))

    for turma in turmas_gerais:

        registradas_planilha = Turma.objects.filter(Ano=ano, CoDisc=turma.CoDisc,
                                                    CodTurma__in=(98,99,97), NroUSP=turma.NroUSP).count()

        registradas_grade_horaria = Turma.objects.filter(Ano=ano, CoDisc=turma.CoDisc,
                                              NroUSP=turma.NroUSP).exclude(CodTurma__in=(98, 99, 97))


        if registradas_grade_horaria.count() > registradas_planilha:

            dif = registradas_grade_horaria.count() - registradas_planilha

            for i in range(dif):
                registradas_grade_horaria[i].delete()

        # if not registradas_planilha:
        #     turma.delete()
        #
        # elif registradas_grade_horaria > 1:



    #RP1
    rp1_tur_p = RP1TurmaPreview.objects.filter(ano=ano)
    rp1_tur = RP1Turma.objects.filter(ano=ano)
    remove_rp(rp1_tur_p, rp1_tur)

    #RP2
    # rp2_tur_p = RP2TurmaPreview.objects.filter(ano=ano)
    # rp2_tur = RP2Turma.objects.filter(ano=ano)
    # remove_rp(rp2_tur_p, rp2_tur)

    #TADI
    tadi_p_tur = TadiTurmaPreview.objects.filter(ano=ano)
    tadi_preview_profs = []

    for t in tadi_p_tur:
        prof = t.professor_si.first() # se for atribuído mais de um prof por turma isso dá errado
        tadi_preview_profs.append(prof)

    dict_tadi_p_profs = contar_professores(tadi_preview_profs)
    tadi_tur = TadiTurma.objects.filter(ano=ano)

    for t in tadi_tur:
        prof = t.professor_si.first()
        if (prof in dict_tadi_p_profs and dict_tadi_p_profs[prof] == 0) or (not prof in dict_tadi_p_profs):
            t.professor_si.remove(prof)

        elif prof in dict_tadi_p_profs:
            dict_tadi_p_profs[prof] -= 1


def contar_professores(lista_professores):
    contagem = {}  # Dicionário para armazenar a contagem de cada professor

    for professor in lista_professores:
        # Se o professor já está no dicionário, incrementa a contagem
        if professor in contagem:
            contagem[professor] += 1
        else:
            # Se o professor não está no dicionário, adiciona com valor 1
            contagem[professor] = 1

    return contagem


def remove_rp(rp_preview, rps):
    lista_prof_preview = []

    for turma in rp_preview:
        profs = turma.professor_si.all()
        for prof in profs:
            lista_prof_preview.append(prof)


    dict_prof_preview = contar_professores(lista_prof_preview)

    for tur in rps:
        profs = tur.professor_si.all()
        for prof in profs:
            if (prof in dict_prof_preview and dict_prof_preview[prof] == 0) or (not prof in dict_prof_preview):
                tur.professor_si.remove(prof)

            elif prof in dict_prof_preview:
                dict_prof_preview[prof] -= 1




#VERSÃO OBSOLETA ABAIXO +++++++++++=
def completa_atrib_obrig_com_hist(discs, ano):
    for disc in discs:
        # sem considerar tadi ou rp1 e rp2
        if disc.CoDisc in ("ACH0041", "ACH0021", "ACH0042"):
            continue

        num_turmas = disc.turma_set.filter(Ano=ano, Eextra="N", CodTurma__in=(99, 98, 97)).count()
        # print("++++++++++++++++++++++++++++++++++++++++=")
        # print(f"disc{disc}; profsMais8{profs_mais_8hrs(disc)}")

        ano = AnoAberto.objects.get(id=1).Ano

        if num_turmas < 3:

            hist = (Turma.objects.filter(CoDisc=disc)
                                    .exclude(NroUSP__in=profs_mais_8hrs(disc))
                                    .values("CoDisc__CoDisc", "NroUSP__NomeProf")
                                    .annotate(vezes_lecionadas=Count("CoDisc__Abreviacao"))
                                    .order_by("-vezes_lecionadas"))

            t_faltando = 3 - num_turmas
            # print(hist)
            # for i in range(0, t_faltando):
            i = 0
            while i < t_faltando:
                try:
                    p = Professor.objects.get(NomeProf=hist[i]['NroUSP__NomeProf'])
                    d = Disciplina.objects.get(CoDisc=hist[i]['CoDisc__CoDisc'])
                    smt = "I" if d.SemestreIdeal in (1, 3, 5, 7) else "P"
                    p_justf = justificativaMenos8Horas.objects.filter(professor=p, semestre_ano=smt, ano=ano)

                    if p_justf:
                        t_faltando += 1

                    else:
                        nova_turma = Turma.objects.create(
                            CoDisc=d,
                            # Código arbitrário
                            CodTurma=99,
                            Ano=ano,
                            NroUSP=p,
                            Eextra="N",
                            SemestreAno=smt,
                            semestre_extra=d.SemestreIdeal
                        )

                    # print(p)
                    # print(d)
                    # print("++++++++++++++++++++++++++++++++++++++++=")

                    i += 1
                except IntegrityError:
                    try:
                        nova_turma = Turma.objects.create(
                            CoDisc=d,
                            # Código arbitrário
                            CodTurma=97,
                            Ano=ano,
                            NroUSP=p,
                            Eextra="N",
                            SemestreAno=smt,
                            semestre_extra=d.SemestreIdeal
                        )
                        i += 1
                    except:
                        if t_faltando + 1 <= hist.count():
                            t_faltando += 1
                        else:
                            print(f"Disciplina {disc.Abreviacao} faltando turmas integridade")

                        i = i + 1
                except IndexError:
                    print(f"Disciplina {disc.Abreviacao} faltando turmas index {hist}")
                    i = t_faltando

# def cria_atribuicao_com_pref(discs, ano_atual):
#
#     for prioridade in range(1, 4):
#         for disc in discs:
#
#             # sem considerar tadi ou rp1 e rp2
#             if disc.CoDisc in ("ACH0041", "ACH0021", "ACH0042"):
#                 continue
#
#             profs_list = profs_mais_8hrs(disc)
#             pref_disc = disc.preferencias_set.filter(AnoProf=ano_atual, nivel=prioridade).exclude(NumProf__in=profs_list)
#
#             num_pref = pref_disc.count()
#             turmas_disc = Turma.objects.filter(CoDisc=disc, Ano=ano_atual, Eextra="N").count()
#
#             #se já houver 3 turmas obrigatórias não é necessário criar outras
#             if turmas_disc >= 3:
#                 continue
#
#             prof_obj = []
#             for pref in pref_disc:
#                 prof_obj.append(pref.NumProf)
#
#             turmas_faltando = 3 - turmas_disc
#
#             if num_pref > turmas_faltando:
#                 #alternativamente dá para carregar na memória tudo e fazer uma busca,
#                 #p melhorar o desempenho
#                 hist_mais_lecionaram = (Turma.objects.filter(CoDisc=disc, NroUSP__in=prof_obj)
#                                         .exclude(NroUSP__in=profs_mais_8hrs(disc))
#                                         .values("CoDisc__CoDisc", "NroUSP__NomeProf")
#                                         .annotate(vezes_lecionadas=Count("CoDisc__Abreviacao"))
#                                         .order_by("-vezes_lecionadas"))
#
#
#                 if hist_mais_lecionaram.count() < turmas_faltando:
#
#                     hist_mais_lecionaram = excecao_profs_sem_hist_matéria(hist_mais_lecionaram, ano_atual, prioridade, disc)
#
#                     for i in range(0, turmas_faltando):
#                         cadastra_turmas_auto(hist_mais_lecionaram[i], False)
#
#                 else:
#
#                     for i in range(0, turmas_faltando):
#                         cadastra_turmas_auto(hist_mais_lecionaram[i], True)
#
#             else:
#                 for pref in pref_disc:
#                     cadastra_turmas_auto(pref, False)
#



# def excecao_profs_sem_hist_matéria(hist_mais_lecionaram, ano_atual, prioridade, disc):
#     # acontece quando profs colocaram para dar a matéria pela primeira vez
#     pref_faltantes = []
#     ids = []
#
#     for hist_dicts in hist_mais_lecionaram:
#         pref_faltante = consulta_pref_por_dict(hist_dicts)
#         ids.append(pref_faltante.id)
#         pref_faltantes.append(pref_faltante)
#
#     new_pref_disc = disc.preferencias_set.filter(AnoProf=ano_atual, nivel=prioridade).exclude(
#         NumProf__in=profs_mais_8hrs(disc)).exclude(id__in=ids)
#
#     return pref_faltantes + list(new_pref_disc)



# def consulta_pref_por_dict(info):
#     ano = AnoAberto.objects.get(id=1).Ano
#     p = Professor.objects.get(NomeProf=info['NroUSP__NomeProf'])
#     d = Disciplina.objects.get(CoDisc=info['CoDisc__CoDisc'])
#     pref = Preferencias.objects.get(NumProf=p, CoDisc=d, AnoProf=ano)
#
#     return pref

# def cadastra_turmas_auto(pref, maior3):
#     ano = AnoAberto.objects.get(id=1).Ano
#     if maior3:
#         pref = consulta_pref_por_dict(pref)
#
#     smt = "P" if pref.Semestre == 2 else "I"
#
#     nova_turma = Turma.objects.create(
#         CoDisc=pref.CoDisc,
#         # Código arbitrário
#         CodTurma=99,
#         Ano=ano,
#         NroUSP=pref.NumProf,
#         Eextra="N",
#         SemestreAno=smt,
#         semestre_extra=pref.CoDisc.SemestreIdeal
#     )